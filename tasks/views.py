from django.shortcuts import render, get_object_or_404, redirect
from .models import Task
from .forms import TaskForm
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from django.contrib.auth import login
from .forms import CustomAuthenticationForm
from .forms import CustomSignUpForm
from django.http import HttpResponse

def export_tasks_to_excel(request):
    tasks = Task.objects.filter(user=request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ['Task Name', 'Description', 'Priority', 'Progress', 'Due Date']
    for col_num, column_title in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=column_title).font = Font(bold=True)

    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task.title)
        ws.cell(row=row_num, column=2, value=task.description)
        ws.cell(row=row_num, column=3, value=task.priority)
        ws.cell(row=row_num, column=4, value=task.progress)
        ws.cell(row=row_num, column=5, value=task.due_date)

    for row_num in range(2, len(tasks) + 2):
        progress_cell = f'D{row_num}'
        ws[f'F{row_num}'] = f'=IF({progress_cell} = 100, "Completed", "In Progress")'

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tasks.xlsx'
    wb.save(response)

    return response

@csrf_exempt
def delete_task(request, task_id):
    try:
        task = Task.objects.get(id=task_id, user=request.user)
        task.delete()
        return JsonResponse({'success': True})
    except Task.DoesNotExist:
        print(f"Task with ID {task_id} does not exist.")
        return JsonResponse({'success': False, 'error': 'Task not found'}, status=404)

def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id, user=request.user)

    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = TaskForm(instance=task)

    return render(request, 'edit_task.html', {'form': form, 'task': task})

def home(request):
    search_query = request.GET.get('search', '')

    if request.user.is_authenticated:
        if search_query:
            tasks = Task.objects.filter(
                Q(title__icontains=search_query) |
                Q(priority__icontains=search_query) |
                Q(description__icontains=search_query),
                user=request.user 
            )
        else:
            tasks = Task.objects.filter(user=request.user)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return render(request, 'partials/tasks_list.html', {'tasks': tasks})

        return render(request, 'home.html', {'tasks': tasks})

    else:
        if request.method == 'POST':
            form = CustomAuthenticationForm(request, data=request.POST)
            if form.is_valid():
                user = form.get_user()
                login(request, user)
                return redirect('home')
        else:
            form = CustomAuthenticationForm(request)

        return render(request, 'home.html', {'form': form, 'tasks': []})

def signup(request):
    if request.method == 'POST':
        form = CustomSignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = CustomSignUpForm()
    
    return render(request, 'signup.html', {'form': form})

def add_task(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            task.save()
            return redirect('home')
    else:
        form = TaskForm()
    tasks = Task.objects.filter(user=request.user)
    return render(request, 'add_task.html', {'tasks': tasks, 'form': form})